import io
from datetime import datetime
from fastapi import APIRouter, HTTPException, Depends, Query
from fastapi.responses import StreamingResponse
from fpdf import FPDF
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from supabase import Client
from dependencies import get_authenticated_supabase

router = APIRouter(prefix="/api/projetos", tags=["Relatorios"])

ETAPAS_OBRA = [
    ("servicos_preliminares", "1. Servicos Preliminares"),
    ("infraestrutura",        "2. Infraestrutura (Fundacao)"),
    ("superestrutura",        "3. Superestrutura (Alvenaria/Lajes)"),
    ("instalacoes",           "4. Instalacoes (Hidraulica/Eletrica)"),
    ("acabamentos",           "5. Acabamentos"),
]

# Larguras das colunas (A4 landscape: 297mm - 2x10mm margens = 277mm)
_COL_COD   = 22
_COL_DESC  = 95
_COL_UNID  = 14
_COL_QTD   = 20
_COL_VUNIT = 30
_COL_BDI   = 14
_COL_VCBDI = 32
_COL_TOTAL = 32  # total = 259mm (sobra para margem interna)
_ROW_H     = 7


class PDFSinapi(FPDF):
    def __init__(self, nome_obra: str, engenheiro: str):
        super().__init__(orientation="L", format="A4")
        self._nome_obra = nome_obra
        self._engenheiro = engenheiro

    def header(self):
        self.set_font("helvetica", "B", 13)
        self.cell(0, 8, "PLANILHA ORCAMENTARIA SINAPI", ln=True, align="C")
        self.set_font("helvetica", "", 9)
        self.cell(0, 5, self._nome_obra, ln=True, align="C")
        self.ln(3)

    def footer(self):
        self.set_y(-14)
        self.set_font("helvetica", "I", 8)
        self.cell(0, 5, f"Pagina {self.page_no()}  |  {self._engenheiro}", align="C")

    def _cabecalho_tabela(self):
        self.set_font("helvetica", "B", 8)
        self.set_fill_color(30, 41, 59)    # slate-800
        self.set_text_color(255, 255, 255)
        self.cell(_COL_COD,  _ROW_H, "Codigo",        border=1, fill=True, align="C")
        self.cell(_COL_DESC, _ROW_H, "Descricao",     border=1, fill=True)
        self.cell(_COL_UNID, _ROW_H, "Unid.",         border=1, fill=True, align="C")
        self.cell(_COL_QTD,  _ROW_H, "Qtd.",          border=1, fill=True, align="C")
        self.cell(_COL_VUNIT,_ROW_H, "V.Unit. (R$)",  border=1, fill=True, align="R")
        self.cell(_COL_BDI,  _ROW_H, "BDI%",          border=1, fill=True, align="C")
        self.cell(_COL_VCBDI,_ROW_H, "V.c/BDI (R$)",  border=1, fill=True, align="R")
        self.cell(_COL_TOTAL,_ROW_H, "Total (R$)",    border=1, fill=True, align="R")
        self.ln()
        self.set_text_color(0, 0, 0)

    def _linha_item(self, item: dict, fator_bdi: float):
        codigo    = item.get("codigo_sinapi") or "MANUAL"
        descricao = item.get("descricao", "")
        unidade   = item.get("unidade", "")
        qtd       = float(item.get("quantidade", 0))
        v_unit    = float(item.get("valor_unitario", 0))
        v_cbdi    = v_unit * fator_bdi
        total     = v_cbdi * qtd

        # Trunca descrição para caber na célula
        descricao_fmt = descricao[:52] if len(descricao) > 52 else descricao

        self.set_font("helvetica", "", 8)
        self.cell(_COL_COD,  _ROW_H, codigo[:12],      border="LR", align="C")
        self.cell(_COL_DESC, _ROW_H, descricao_fmt,    border="LR")
        self.cell(_COL_UNID, _ROW_H, unidade,          border="LR", align="C")
        self.cell(_COL_QTD,  _ROW_H, f"{qtd:,.3f}",   border="LR", align="R")
        self.cell(_COL_VUNIT,_ROW_H, f"{v_unit:,.2f}", border="LR", align="R")
        self.cell(_COL_BDI,  _ROW_H, "",               border="LR", align="C")  # preenchido por etapa
        self.cell(_COL_VCBDI,_ROW_H, f"{v_cbdi:,.2f}", border="LR", align="R")
        self.cell(_COL_TOTAL,_ROW_H, f"{total:,.2f}",  border="LR", align="R")
        self.ln()
        return total

    def _linha_subtotal(self, label: str, subtotal: float):
        largura_esq = _COL_COD + _COL_DESC + _COL_UNID + _COL_QTD + _COL_VUNIT + _COL_BDI + _COL_VCBDI
        self.set_font("helvetica", "B", 8)
        self.set_fill_color(241, 245, 249)  # slate-100
        self.cell(largura_esq, _ROW_H, f"  Subtotal {label}", border=1, fill=True, align="L")
        self.cell(_COL_TOTAL,  _ROW_H, f"R$ {subtotal:,.2f}", border=1, fill=True, align="R")
        self.ln()

    def _linha_total_geral(self, total: float):
        largura_esq = _COL_COD + _COL_DESC + _COL_UNID + _COL_QTD + _COL_VUNIT + _COL_BDI + _COL_VCBDI
        self.set_font("helvetica", "B", 10)
        self.set_fill_color(30, 41, 59)
        self.set_text_color(255, 255, 255)
        self.cell(largura_esq, 9, "  TOTAL GERAL (com BDI)", border=1, fill=True, align="L")
        self.cell(_COL_TOTAL,  9, f"R$ {total:,.2f}",        border=1, fill=True, align="R")
        self.set_text_color(0, 0, 0)
        self.ln()

class PDFProposta(FPDF):
    def header(self):
        self.set_font('helvetica', 'B', 16)
        self.cell(0, 10, 'Proposta Comercial - Orcamento de Obra', ln=True, align='C')
        self.ln(5)

    def footer(self):
        self.set_y(-30)
        self.set_font('helvetica', 'I', 10)
        
        # Linha para assinatura centralizada
        largura_linha = 80
        margem_esq = (self.w - largura_linha) / 2
        self.line(margem_esq, self.get_y(), margem_esq + largura_linha, self.get_y())
        
        self.cell(0, 10, 'Engenheiro Responsavel', ln=True, align='C')
        self.set_font('helvetica', 'I', 8)
        self.cell(0, 10, f'Pagina {self.page_no()}', align='C')

@router.get("/{projeto_id}/pdf-comercial")
def gerar_pdf_comercial(
    projeto_id: str,
    supabase_client: Client = Depends(get_authenticated_supabase)
):
    """
    Gera PDF de proposta comercial com BDI embutido (cliente não vê custo base).
    Roda como `def` (sync) para delegação automática ao thread pool. REF: TODO.md P1.2
    """
    try:
        # 1. Busca os dados do Projeto na tabela unificada
        proj_res = supabase_client.table("projetos_clientes").select("*").eq("id", projeto_id).single().execute()
        if not proj_res.data:
            raise HTTPException(status_code=404, detail="Projeto não encontrado.")
        projeto = proj_res.data

        # 2. Busca a lista de itens vinculados a este Projeto
        itens_res = supabase_client.table("orcamento_itens").select("*").eq("projeto_id", projeto_id).execute()
        itens = itens_res.data or []

        # 3. Lógica Matemática do BDI invisível
        bdi_percentual = float(projeto.get("bdi_padrao") or 0.0)
        fator_bdi = 1 + (bdi_percentual / 100)

        # 4. Desenha o PDF
        pdf = PDFProposta()
        pdf.add_page()

        # Cabeçalho de Informações
        pdf.set_font("helvetica", size=12)
        pdf.cell(0, 8, txt=f"Obra: {projeto.get('titulo_projeto', 'Não informada')}", ln=True)
        cliente_nome = projeto.get('cliente_nome')
        pdf.cell(0, 8, txt=f"Cliente: {cliente_nome if cliente_nome else 'Não informado'}", ln=True)
        data_atual = datetime.now().strftime("%d/%m/%Y")
        pdf.cell(0, 8, txt=f"Data: {data_atual}", ln=True)
        pdf.ln(10)

        # Cabeçalho da Tabela
        pdf.set_font("helvetica", 'B', 10)
        pdf.cell(85, 8, "Item/Descricao", border=1)
        pdf.cell(15, 8, "Unid.", border=1, align='C')
        pdf.cell(15, 8, "Qtd.", border=1, align='C')
        pdf.cell(35, 8, "Valor Unit. (R$)", border=1, align='C')
        pdf.cell(35, 8, "Total (R$)", border=1, align='C')
        pdf.ln()

        # Corpo da Tabela
        pdf.set_font("helvetica", '', 10)
        valor_total_proposta = 0

        for item in itens:
            descricao = item.get("descricao", "Item sem nome")
            unidade = item.get("unidade", "un")
            quantidade = float(item.get("quantidade", 0))
            valor_custo = float(item.get("valor_unitario", 0))

            # Aplica o BDI na venda: Regra Ouro (Cliente final não vê o custo puro nem a sigla BDI)
            valor_venda_unitario = valor_custo * fator_bdi
            valor_venda_total = valor_venda_unitario * quantidade
            valor_total_proposta += valor_venda_total

            # Corta strings muito longas para não quebrar o layout da tabela (máx ~45 caracteres)
            pdf.cell(85, 8, descricao[:45], border=1)
            pdf.cell(15, 8, unidade, border=1, align='C')
            pdf.cell(15, 8, f"{quantidade:.2f}", border=1, align='C')
            pdf.cell(35, 8, f"{valor_venda_unitario:.2f}", border=1, align='R')
            pdf.cell(35, 8, f"{valor_venda_total:.2f}", border=1, align='R')
            pdf.ln()

        pdf.ln(5)

        # Valor Total da Proposta (Rodapé da Tabela)
        pdf.set_font("helvetica", 'B', 12)
        pdf.cell(150, 10, "Valor Total da Proposta:", border=0, align='R')
        pdf.cell(35, 10, f"R$ {valor_total_proposta:.2f}", border=0, align='R')

        # Output para Byte String em Memória (Não salva no disco)
        pdf_bytes = pdf.output()
        buffer = io.BytesIO(pdf_bytes)
        buffer.seek(0)

        # Retorna o arquivo binário anexado
        headers = {
            "Content-Disposition": f'attachment; filename="proposta_comercial_{projeto_id}.pdf"'
        }
        return StreamingResponse(buffer, media_type="application/pdf", headers=headers)
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(status_code=500, detail="Erro ao gerar a proposta comercial. Tente novamente.")


@router.get("/{projeto_id}/exportar-sinapi")
def exportar_planilha_sinapi(
    projeto_id: str,
    formato: str = Query("pdf", pattern="^(pdf|xlsx)$"),
    supabase_client: Client = Depends(get_authenticated_supabase)
):
    """
    Gera planilha orçamentária SINAPI em PDF (A4 paisagem) ou Excel (.xlsx).
    Itens agrupados por etapa, BDI aplicado, subtotais e total geral.
    Roda como `def` (sync) — delega ao thread pool automaticamente.
    """
    try:
        proj_res = supabase_client.table("projetos_clientes").select("*").eq("id", projeto_id).single().execute()
        if not proj_res.data:
            raise HTTPException(status_code=404, detail="Projeto não encontrado.")
        projeto = proj_res.data

        itens_res = supabase_client.table("orcamento_itens").select("*").eq("projeto_id", projeto_id).execute()
        itens = itens_res.data or []

        if not itens:
            raise HTTPException(status_code=422, detail="O projeto não possui itens para exportar.")

        # --- Dados comuns ---
        nome_obra    = projeto.get("titulo_projeto") or "Obra sem nome"
        cliente_nome = projeto.get("cliente_nome") or "Não informado"
        uf_obra      = projeto.get("uf_obra") or "-"
        mes_ref      = projeto.get("sinapi_mes_ano") or "-"
        desonerado   = "Sim" if projeto.get("sinapi_desonerado") else "Nao"
        bdi_perc     = float(projeto.get("bdi_padrao") or 0.0)
        fator_bdi    = 1 + (bdi_perc / 100)
        data_geracao = datetime.now().strftime("%d/%m/%Y")
        slug_nome    = nome_obra[:30].replace(" ", "_")

        # Agrupa itens por etapa preservando a ordem de ETAPAS_OBRA
        etapa_map = {chave: [] for chave, _ in ETAPAS_OBRA}
        sem_etapa = []
        for item in itens:
            chave = item.get("etapa_obra") or ""
            if chave in etapa_map:
                etapa_map[chave].append(item)
            else:
                sem_etapa.append(item)

        grupos_ordenados = [(chave, label, etapa_map[chave]) for chave, label in ETAPAS_OBRA if etapa_map[chave]]
        if sem_etapa:
            grupos_ordenados.append(("outros", "Outros", sem_etapa))

        # ------------------------------------------------------------------ #
        #  FORMATO PDF                                                         #
        # ------------------------------------------------------------------ #
        if formato == "pdf":
            pdf = PDFSinapi(
                nome_obra=f"Obra: {nome_obra} | Cliente: {cliente_nome}",
                engenheiro=f"Gerado em {data_geracao}"
            )
            pdf.set_margins(10, 10, 10)
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()

            pdf.set_font("helvetica", "B", 9)
            pdf.cell(45, 6, "Obra:",               border=0)
            pdf.set_font("helvetica", "", 9)
            pdf.cell(0,  6, nome_obra,             border=0, ln=True)

            pdf.set_font("helvetica", "B", 9)
            pdf.cell(45, 6, "Cliente:",            border=0)
            pdf.set_font("helvetica", "", 9)
            pdf.cell(0,  6, cliente_nome,          border=0, ln=True)

            pdf.set_font("helvetica", "B", 9)
            pdf.cell(45, 6, "UF da Obra:",         border=0)
            pdf.set_font("helvetica", "", 9)
            pdf.cell(60, 6, uf_obra,               border=0)
            pdf.set_font("helvetica", "B", 9)
            pdf.cell(45, 6, "Mes de Referencia:",  border=0)
            pdf.set_font("helvetica", "", 9)
            pdf.cell(0,  6, mes_ref,               border=0, ln=True)

            pdf.set_font("helvetica", "B", 9)
            pdf.cell(45, 6, "Desonerado:",         border=0)
            pdf.set_font("helvetica", "", 9)
            pdf.cell(60, 6, desonerado,            border=0)
            pdf.set_font("helvetica", "B", 9)
            pdf.cell(45, 6, "BDI:",                border=0)
            pdf.set_font("helvetica", "", 9)
            pdf.cell(0,  6, f"{bdi_perc:.1f}%",   border=0, ln=True)

            pdf.ln(4)
            pdf._cabecalho_tabela()
            total_geral = 0.0

            for _chave, label, grupo in grupos_ordenados:
                largura_total = _COL_COD + _COL_DESC + _COL_UNID + _COL_QTD + _COL_VUNIT + _COL_BDI + _COL_VCBDI + _COL_TOTAL
                pdf.set_font("helvetica", "B", 8)
                pdf.set_fill_color(226, 232, 240)
                pdf.cell(largura_total, _ROW_H, f"  {label}", border=1, fill=True, ln=True)

                subtotal = 0.0
                for item in grupo:
                    if pdf.get_y() > pdf.h - 30:
                        pdf.add_page()
                        pdf._cabecalho_tabela()
                    subtotal += pdf._linha_item(item, fator_bdi)

                pdf._linha_subtotal(label, subtotal)
                total_geral += subtotal

            pdf.ln(3)
            pdf._linha_total_geral(total_geral)

            buffer = io.BytesIO(pdf.output())
            buffer.seek(0)
            nome_arquivo = f"planilha_sinapi_{slug_nome}_{mes_ref}.pdf"
            headers = {"Content-Disposition": f'attachment; filename="{nome_arquivo}"'}
            return StreamingResponse(buffer, media_type="application/pdf", headers=headers)

        # ------------------------------------------------------------------ #
        #  FORMATO XLSX                                                        #
        # ------------------------------------------------------------------ #
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Planilha SINAPI"

        # Estilos reutilizáveis
        _thin = Side(style="thin", color="CCCCCC")
        _border_all = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

        def _fill(hex_color: str) -> PatternFill:
            return PatternFill("solid", fgColor=hex_color)

        def _font(bold=False, color="000000", size=10):
            return Font(bold=bold, color=color, size=size, name="Calibri")

        # ── Bloco de informações do projeto (linhas 1-6) ──
        info_campos = [
            ("Obra",             nome_obra),
            ("Cliente",          cliente_nome),
            ("UF da Obra",       uf_obra),
            ("Mês de Referência", mes_ref),
            ("Desonerado",       desonerado),
            ("BDI",              f"{bdi_perc:.1f}%"),
        ]
        for i, (campo, valor) in enumerate(info_campos, start=1):
            ws.cell(i, 1, campo).font  = _font(bold=True)
            ws.cell(i, 2, valor).font  = _font()
            ws.cell(i, 1).fill = _fill("F1F5F9")

        linha_atual = len(info_campos) + 2  # linha em branco de separação

        # ── Cabeçalho das colunas ──
        cabecalhos = [
            ("Código SINAPI", 16),
            ("Descrição",     60),
            ("Unidade",       10),
            ("Quantidade",    14),
            ("V. Unit. (R$)", 16),
            ("BDI (%)",       10),
            ("V. c/BDI (R$)", 16),
            ("Total (R$)",    16),
        ]
        for col_idx, (titulo, largura) in enumerate(cabecalhos, start=1):
            cell = ws.cell(linha_atual, col_idx, titulo)
            cell.font    = _font(bold=True, color="FFFFFF")
            cell.fill    = _fill("1E293B")   # slate-800
            cell.border  = _border_all
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
            ws.column_dimensions[get_column_letter(col_idx)].width = largura

        ws.row_dimensions[linha_atual].height = 20
        linha_atual += 1

        total_geral = 0.0

        for _chave, label, grupo in grupos_ordenados:
            # Linha de cabeçalho da etapa (mesclada nas 8 colunas)
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=8)
            cell_etapa = ws.cell(linha_atual, 1, f"  {label}")
            cell_etapa.font      = _font(bold=True, size=10)
            cell_etapa.fill      = _fill("E2E8F0")   # slate-200
            cell_etapa.border    = _border_all
            cell_etapa.alignment = Alignment(vertical="center")
            ws.row_dimensions[linha_atual].height = 18
            linha_atual += 1

            subtotal = 0.0
            for item in grupo:
                qtd    = float(item.get("quantidade", 0))
                v_unit = float(item.get("valor_unitario", 0))
                v_cbdi = v_unit * fator_bdi
                total  = v_cbdi * qtd
                subtotal += total

                valores = [
                    item.get("codigo_sinapi") or "MANUAL",
                    item.get("descricao", ""),
                    item.get("unidade", ""),
                    qtd,
                    v_unit,
                    bdi_perc,
                    v_cbdi,
                    total,
                ]
                for col_idx, valor in enumerate(valores, start=1):
                    cell = ws.cell(linha_atual, col_idx, valor)
                    cell.border = _border_all
                    cell.font   = _font(size=9)
                    if col_idx in (4, 5, 6, 7, 8):  # colunas numéricas
                        cell.alignment = Alignment(horizontal="right")
                        if col_idx == 4:
                            cell.number_format = "#,##0.000"
                        elif col_idx == 6:
                            cell.number_format = "0.00"
                        else:
                            cell.number_format = '#,##0.00'
                ws.row_dimensions[linha_atual].height = 15
                linha_atual += 1

            # Linha de subtotal da etapa
            ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=7)
            cell_sub_label = ws.cell(linha_atual, 1, f"  Subtotal — {label}")
            cell_sub_label.font      = _font(bold=True, size=9)
            cell_sub_label.fill      = _fill("F1F5F9")
            cell_sub_label.border    = _border_all
            cell_sub_label.alignment = Alignment(vertical="center")

            cell_sub_val = ws.cell(linha_atual, 8, subtotal)
            cell_sub_val.font          = _font(bold=True, size=9)
            cell_sub_val.fill          = _fill("F1F5F9")
            cell_sub_val.border        = _border_all
            cell_sub_val.number_format = '#,##0.00'
            cell_sub_val.alignment     = Alignment(horizontal="right")
            ws.row_dimensions[linha_atual].height = 16
            linha_atual += 1
            total_geral += subtotal

        # Linha em branco
        linha_atual += 1

        # Linha de total geral
        ws.merge_cells(start_row=linha_atual, start_column=1, end_row=linha_atual, end_column=7)
        cell_tot_label = ws.cell(linha_atual, 1, "  TOTAL GERAL (com BDI)")
        cell_tot_label.font      = _font(bold=True, color="FFFFFF", size=11)
        cell_tot_label.fill      = _fill("1E293B")
        cell_tot_label.border    = _border_all
        cell_tot_label.alignment = Alignment(vertical="center")

        cell_tot_val = ws.cell(linha_atual, 8, total_geral)
        cell_tot_val.font          = _font(bold=True, color="FFFFFF", size=11)
        cell_tot_val.fill          = _fill("1E293B")
        cell_tot_val.border        = _border_all
        cell_tot_val.number_format = '"R$ "#,##0.00'
        cell_tot_val.alignment     = Alignment(horizontal="right")
        ws.row_dimensions[linha_atual].height = 22

        # Congela cabeçalho das colunas
        ws.freeze_panes = ws.cell(len(info_campos) + 3, 1)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        nome_arquivo = f"planilha_sinapi_{slug_nome}_{mes_ref}.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{nome_arquivo}"',
        }
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao gerar a planilha SINAPI: {str(e)}")